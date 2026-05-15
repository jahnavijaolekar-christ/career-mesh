from flask import send_from_directory
from flask import Flask, render_template, request, redirect, url_for, jsonify
from flask_cors import CORS
import openpyxl
import os
from datetime import datetime
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app)

# ── Excel DB setup ──────────────────────────────────────────
DB_PATH = "database/applications.xlsx"
UPLOAD_FOLDER = "database/resumes"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def init_db():
    if not os.path.exists(DB_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append([
            "ID", "Name", "Email", "Phone",
            "Job Title", "Experience", "Cover Letter",
            "Resume", "Skills Match %", "JD Match %",
            "Status", "Interview Slot", "Applied On"
        ])
        wb.save(DB_PATH)

def extract_text_from_pdf(filepath):
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text.strip()
    except Exception as e:
        print(f"PDF extraction error: {e}")
        return ""

def calculate_ats_scores(resume_text, job_title, required_skills, jd_text):
    if not resume_text:
        return 0, 0
    resume_lower = resume_text.lower()
    skills_list    = [s.strip().lower() for s in required_skills.split(",") if s.strip()]
    matched_skills = [s for s in skills_list if s in resume_lower]
    skills_match   = round((len(matched_skills) / len(skills_list)) * 100) if skills_list else 0
    stop_words = {
        'the','and','for','with','that','this','are','will','have',
        'from','your','our','you','not','all','been','has','its',
        'into','more','also','their','they','when','than','about',
        'what','which','who','can','each','other','such','these',
        'those','then','some','would','make','like','him','time',
        'very','just','know','take','year','good','come','could',
        'well','even','want','look','use','see','may','day','way'
    }
    jd_keywords = set(
        w.lower().strip('.,;:()[]') for w in jd_text.split()
        if len(w) > 3
        and w.lower().strip('.,;:()[]') not in stop_words
        and w.isalpha()
    )
    matched_jd = [w for w in jd_keywords if w in resume_lower]
    jd_match   = min(round((len(matched_jd) / len(jd_keywords)) * 120), 100) if jd_keywords else 0
    return skills_match, jd_match

def get_all_applications():
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb.active
    apps = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            apps.append({
                "id":             row[0],
                "name":           row[1],
                "email":          row[2],
                "phone":          row[3],
                "job_title":      row[4],
                "experience":     row[5],
                "cover_letter":   row[6],
                "resume":         row[7]  if len(row) > 7  and row[7]  is not None else "",
                "skills_match":   row[8]  if len(row) > 8  and row[8]  is not None else 0,
                "jd_match":       row[9]  if len(row) > 9  and row[9]  is not None else 0,
                "status":         row[10] if len(row) > 10 and row[10] is not None else "Under Review",
                "interview_slot": row[11] if len(row) > 11 and row[11] is not None else "",
                "applied_on":     row[12] if len(row) > 12 and row[12] is not None else "",
            })
    return apps

def save_application(data, resume_file=None):
    wb     = openpyxl.load_workbook(DB_PATH)
    ws     = wb.active
    new_id = ws.max_row
    resume_filename = ""
    skills_match    = 0
    jd_match        = 0
    if resume_file and resume_file.filename:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        filename        = secure_filename(resume_file.filename)
        resume_filename = f"{new_id}_{filename}"
        resume_path     = os.path.join(UPLOAD_FOLDER, resume_filename)
        resume_file.save(resume_path)
        resume_text = extract_text_from_pdf(resume_path)
        if resume_text:
            job             = next((j for j in JOBS if j["title"] == data["job_title"]), None)
            jd_text         = job.get("jd", job.get("description", "")) if job else ""
            required_skills = job.get("skills", "") if job else ""
            skills_match, jd_match = calculate_ats_scores(
                resume_text, data["job_title"], required_skills, jd_text
            )
    status = "Shortlisted" if skills_match >= 70 else "Under Review"
    ws.append([
        new_id,
        data["name"],
        data["email"],
        data["phone"],
        data["job_title"],
        data["experience"],
        data["cover_letter"],
        resume_filename,
        skills_match,
        jd_match,
        status,
        "",
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ])
    wb.save(DB_PATH)
    return skills_match, status

# ── Job listings (populated dynamically by Copilot Agent) ────
JOBS = []

# ── Routes ───────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html", jobs=JOBS)

@app.route("/apply/<int:job_id>", methods=["GET", "POST"])
def apply(job_id):
    job = next((j for j in JOBS if j["id"] == job_id), None)
    if not job:
        return "Job not found", 404
    if request.method == "POST":
        try:
            resume_file = request.files.get("resume")
            save_application({
                "name":         request.form.get("name", ""),
                "email":        request.form.get("email", ""),
                "phone":        request.form.get("phone", ""),
                "job_title":    job["title"],
                "experience":   request.form.get("experience", ""),
                "cover_letter": request.form.get("cover_letter", ""),
            }, resume_file)
            return redirect(url_for("success"))
        except Exception as e:
            return f"Error submitting application: {e}", 500
    return render_template("apply.html", job=job)

@app.route("/success")
def success():
    return render_template("success.html")

@app.route("/dashboard")
def dashboard():
    all_apps = get_all_applications()
    shortlisted  = sorted(
        [a for a in all_apps if a["status"] == "Shortlisted"],
        key=lambda x: x["skills_match"], reverse=True
    )
    under_review = sorted(
        [a for a in all_apps if a["status"] != "Shortlisted"],
        key=lambda x: x["skills_match"], reverse=True
    )
    return render_template("dashboard.html",
                           shortlisted=shortlisted,
                           under_review=under_review,
                           total=len(all_apps))

@app.route("/job/<int:job_id>")
def view_jd(job_id):
    job = next((j for j in JOBS if j["id"] == job_id), None)
    if not job:
        return "Job not found", 404
    return render_template("jd.html", job=job)

@app.route("/resume/<filename>")
def serve_resume(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route("/api/jobs")
def api_jobs():
    return jsonify(JOBS)

@app.route("/api/applications")
def api_applications():
    return jsonify(get_all_applications())

@app.route("/api/post-job-from-agent", methods=["POST"])
def post_job_from_agent():
    data = request.json
    required_fields = ["title", "company", "location", "type", "description"]
    for field in required_fields:
        if not data.get(field):
            return jsonify({"error": f"Missing field: {field}"}), 400
    new_job = {
        "id":          len(JOBS) + 1,
        "title":       data["title"],
        "company":     data["company"],
        "location":    data["location"],
        "type":        data["type"],
        "description": data["description"][:120] + "..." if len(data["description"]) > 120 else data["description"],
        "skills":      data.get("skills", ""),
        "posted_by":   data.get("posted_by", "Copilot Agent"),
        "jd":          data["description"],
    }
    JOBS.append(new_job)
    return jsonify({
        "success": True,
        "message": f"Job '{new_job['title']}' posted successfully!",
        "job":     new_job
    })

if __name__ == "__main__":
    init_db()
    app.run(debug=False)
